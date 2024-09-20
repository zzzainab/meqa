import os
from tqdm import tqdm
import logging
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from openpyxl import load_workbook
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from google.oauth2 import service_account
from langchain_openai import ChatOpenAI
from langchain.prompts import ChatPromptTemplate
from langchain.schema.runnable import RunnablePassthrough
from dotenv import load_dotenv
import time
import random
import logging
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

# Configure the logger
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Suppress oauth2client and googleapiclient.discovery messages
logging.getLogger('oauth2client').setLevel(logging.ERROR)
logging.getLogger('googleapiclient.discovery').setLevel(logging.ERROR)
logging.getLogger('googleapiclient.discovery_cache').setLevel(logging.ERROR)

# Load environment variables from .env file
load_dotenv()

# Set your OpenAI API key
os.environ["OPENAI_API_KEY"] = os.getenv("OPENAI_API_KEY")


# Excel file path
excel_file = os.path.join(os.path.dirname(__file__), 'criteria-excel.xlsx')

# Google Spreadsheet ID
SPREADSHEET_ID = os.getenv("GOOGLE_SPREADSHEET_ID_CYBERSECEVALCOPY")

# Load the Excel workbook
wb = load_workbook(filename=excel_file)
sheet = wb.active

# Initialize the GPT-4o model
llm = ChatOpenAI(model_name="gpt-4o")

# Function to get Google Sheets service
def get_google_sheets_service():
    path_service_account = os.path.join(os.path.dirname(__file__), 'apart-cyberbenchmarks-705df8a67117.json')
    creds = service_account.Credentials.from_service_account_file(
        path_service_account,
        scopes=['https://www.googleapis.com/auth/spreadsheets']
    )
    return build('sheets', 'v4', credentials=creds)

# Function to read from Google Sheets
def read_google_sheets(range_name):
    service = get_google_sheets_service()
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=SPREADSHEET_ID, range=range_name).execute()
    return result.get('values', [])

# Function to get few-shot examples from Google Sheets
def get_few_shot_examples(excel_column, google_sheet_column):
    examples = ""
    score_range = f'{google_sheet_column}14:{google_sheet_column}20'
    comment_range = f'{google_sheet_column}31:{google_sheet_column}37'  # 17 rows below the scores
    scores = read_google_sheets(score_range)
    comments = read_google_sheets(comment_range)
    
    for row, (score, comment) in enumerate(zip(scores, comments), start=5):
        if score and comment and score[0] != "TBD" and comment[0] != "TBD":
            criteria = sheet.cell(row=4, column=excel_column).value
            explanation = sheet.cell(row=row, column=excel_column).value
            examples += f"1)Criteria: {criteria}\n\n2)Original Explanation: {explanation}\n\n3)Score: {score[0]}\n\n4)Final comment: {comment[0]}\n\n\n ############################## \n\n\n"
    
    return examples

# Create prompt template
prompt_template = """For a given benchmark we have to evaluate this criteria: '''1)Criteria: {criteria}'''

Here is the explanation about how well the benchmark fits this criteria: '''2)Original Explanation: {explanation}'''

Based on this information, come up with a score from 1 to 5 (where 5 is the best fit) or N/A if the criteria does not apply to the benchmark. You should also provide a comment explaining your score and for that inspire yourself in the following examples. In the examples the same criteria is evaluated and you can 

Here are a few examples to help you understand how to score this criteria:
'''
{few_shot_examples}
'''

Provide your response in the following format:
3)Score: [Your score]
4)Final Comment: [Your comment explaining the score]

"""

prompt = ChatPromptTemplate.from_template(prompt_template)


# Main evaluation function
def evaluate_benchmark(criteria, explanation, excel_column, google_sheet_column):
    few_shot_examples = get_few_shot_examples(excel_column, google_sheet_column)

    chain = (
        {"criteria": RunnablePassthrough(), "explanation": RunnablePassthrough(), "few_shot_examples": RunnablePassthrough()}
        | prompt
        | llm
    )
    
    result = chain.invoke({"criteria": criteria, "explanation": explanation, "few_shot_examples": few_shot_examples})

    
    # Extract score and comment from the result
    lines = result.content.replace('*', '').split('\n')
    score = lines[0].split(': ')[1]
    # Assuming 'lines' is a list of strings representing the lines in the file
    index = 1
    comment_line = lines[index]
    # Iterate until a non-empty line that does start with "4)Final Comment: " is found
    # Assuming 'lines' is a list of strings representing the lines in the file
    index = 1
    comment_line = lines[index]
    
    # Iterate until a line that starts with "4)Final Comment: " is found
    while "Final Comment: " not in comment_line:
        index += 1
        if index >= len(lines):
            break
        comment_line = lines[index]
    comment = comment_line.split(': ')[1]
    
    return score, comment, criteria, explanation, few_shot_examples


def exponential_backoff(attempt):
    return min(32, 2 ** attempt) + random.random()

def read_cell_value(cell_range, max_retries=5):
    for attempt in range(max_retries):
        try:
            service = get_google_sheets_service()
            sheet = service.spreadsheets()
            result = sheet.values().get(spreadsheetId=SPREADSHEET_ID, range=cell_range).execute()
            values = result.get('values', [])
            if values:
                return values[0][0]  # Return the value of the cell
            else:
                return None  # Cell is empty
        except HttpError as e:
            if e.resp.status == 429:  # Rate limit exceeded
                wait_time = exponential_backoff(attempt)
                logger.warning(f"Rate limit exceeded. Retrying in {wait_time:.2f} seconds.")
                time.sleep(wait_time)
            else:
                raise
    logger.error(f"Failed to read cell {cell_range} after {max_retries} attempts.")
    return None

def write_google_sheets(range_name, values, max_retries=5):
    for attempt in range(max_retries):
        try:
            service = get_google_sheets_service()
            sheet = service.spreadsheets()
            body = {'values': values}
            result = sheet.values().update(
                spreadsheetId=SPREADSHEET_ID,
                range=range_name,
                valueInputOption='USER_ENTERED',
                body=body
            ).execute()
            logger.info(f"Successfully updated range {range_name}")
            return result
        except HttpError as e:
            if e.resp.status == 429:  # Rate limit exceeded
                wait_time = exponential_backoff(attempt)
                logger.warning(f"Rate limit exceeded. Retrying in {wait_time:.2f} seconds.")
                time.sleep(wait_time)
            else:
                raise
    logger.error(f"Failed to write to range {range_name} after {max_retries} attempts.")
    return None
def is_cell_tbd(cell_range):
    value = read_cell_value(cell_range)
    return value == "TBD"


def excel_column_to_letter(column_number):
    """Convert Excel column number to column letter(s)."""
    column_letter = ''
    while column_number > 0:
        column_number -= 1
        column_letter = chr(65 + column_number % 26) + column_letter
        column_number //= 26
    return column_letter

# Main execution
for row in tqdm([13, 15], desc="Processing rows"):  # Rows 5 to 11 in Excel
    for col in range(3, 47):  # Columns C to AU in Excel
        criteria = sheet.cell(row=4, column=col).value + ".\n Note: " + sheet.cell(row=row-1, column=col).value
        # criteria = sheet.cell(row=4, column=col).value
        explanation = sheet.cell(row=row, column=col).value
        
        if criteria and explanation:
            google_sheet_col = excel_column_to_letter(col-1)  # Convert to A, B, C, ..., AA, AB, etc.
            if row == 13:
                google_sheet_score_row = 10  # Row 10 in Google Sheets
            elif row == 15:
                google_sheet_score_row = 11 # Row 12 in Google Sheets
            # google_sheet_score_row = row + 9  # Rows 14 to 20 in Google Sheets
            google_sheet_comment_row = google_sheet_score_row + 17  # 17 rows below the score
            
            score_cell = f'{google_sheet_col}{google_sheet_score_row}'
            comment_cell = f'{google_sheet_col}{google_sheet_comment_row}'
            
            if is_cell_tbd(score_cell):
                score, comment, criteria, explanation, few_shot_examples = evaluate_benchmark(criteria, explanation, col, google_sheet_col)
                
                if write_google_sheets(score_cell, [[score]]) is not None:
                    logger.info(f"Updated score in cell {score_cell}")
                else:
                    logger.error(f"Failed to update score in cell {score_cell}")
            
                # Write comment to Google Sheets if it's TBD
                if is_cell_tbd(comment_cell):
                    if write_google_sheets(comment_cell, [[comment]]) is not None:
                        logger.info(f"Updated comment in cell {comment_cell}")
                    else:
                        logger.error(f"Failed to update comment in cell {comment_cell}")
            
                # Add a small delay between operations to help with rate limiting
                time.sleep(1)

            else:
                # logger.info(f"Skipped cells {score_cell} and {comment_cell} as they are not TBD")
                pass

print("Evaluation complete. Google Sheets updated.")